from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "src" / "data" / "slotLayouts.js"
OUTPUT = ROOT / "stage_grids.xlsx"
GRID_COLS = 9
GRID_ROWS = 12


def parse_stage_layouts() -> list[dict[str, list[list[int]]]]:
    source = SOURCE.read_text(encoding="utf-8")
    match = re.search(r"const STAGE_LAYOUTS = \[(.*)\];\s+const stages", source, re.S)
    if not match:
        raise RuntimeError("STAGE_LAYOUTS block not found")
    block = match.group(1)
    object_blocks = re.findall(r"\{\s*path:\s*\[(.*?)\],\s*slots:\s*\[(.*?)\],\s*\}", block, re.S)
    stages = []
    for path_block, slot_block in object_blocks:
      path = [[int(x), int(y)] for x, y in re.findall(r"\[(-?\d+),\s*(-?\d+)\]", path_block)]
      slots = [[int(x), int(y)] for x, y in re.findall(r"\[(\d+),\s*(\d+)\]", slot_block)]
      stages.append({"path": path, "slots": slots})
    return stages


def expand_path_points(path: list[list[int]]) -> set[tuple[int, int]]:
    points: set[tuple[int, int]] = set()
    for start, end in zip(path, path[1:]):
        x1, y1 = start
        x2, y2 = end
        points.add((x1, y1))
        if x1 == x2:
            step = 1 if y2 >= y1 else -1
            for y in range(y1, y2 + step, step):
                points.add((x1, y))
        elif y1 == y2:
            step = 1 if x2 >= x1 else -1
            for x in range(x1, x2 + step, step):
                points.add((x, y1))
        else:
            raise ValueError(f"Path segment must be axis-aligned: {(x1, y1)} -> {(x2, y2)}")
    if path:
        points.add(tuple(path[-1]))
    return points


def build_workbook(stage_layouts: list[dict[str, list[list[int]]]]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    fills = {
        "empty": PatternFill("solid", fgColor="0E1421"),
        "path": PatternFill("solid", fgColor="7A3642"),
        "slot": PatternFill("solid", fgColor="1A3A5F"),
        "both": PatternFill("solid", fgColor="7A5A20"),
    }
    font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
    axis_font = Font(name="Consolas", size=10, bold=True, color="D6E2F5")
    border = Border(
        left=Side(style="thin", color="31425C"),
        right=Side(style="thin", color="31425C"),
        top=Side(style="thin", color="31425C"),
        bottom=Side(style="thin", color="31425C"),
    )

    for index, stage in enumerate(stage_layouts, start=1):
        ws = wb.create_sheet(f"Stage {index}")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B2"

        for col in range(1, GRID_COLS + 2):
            ws.column_dimensions[chr(64 + col)].width = 5.2
        for row in range(1, GRID_ROWS + 2):
            ws.row_dimensions[row].height = 24

        for x in range(GRID_COLS):
            cell = ws.cell(row=1, column=x + 2, value=x)
            cell.font = axis_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for y in range(GRID_ROWS):
            cell = ws.cell(row=y + 2, column=1, value=y)
            cell.font = axis_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        path_points = expand_path_points(stage["path"])
        slot_points = {tuple(point) for point in stage["slots"]}
        start_point = tuple(stage["path"][0]) if stage["path"] else None
        end_point = tuple(stage["path"][-1]) if stage["path"] else None

        for y in range(GRID_ROWS):
            for x in range(GRID_COLS):
                point = (x, y)
                value = "."
                fill = fills["empty"]
                if point == start_point:
                    value = "A"
                    fill = fills["path"]
                elif point == end_point:
                    value = "Z"
                    fill = fills["path"]
                elif point in path_points and point in slot_points:
                    value = "B"
                    fill = fills["both"]
                elif point in path_points:
                    value = "P"
                    fill = fills["path"]
                elif point in slot_points:
                    value = "S"
                    fill = fills["slot"]
                cell = ws.cell(row=y + 2, column=x + 2, value=value)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        ws["L2"] = "Legend"
        ws["L2"].font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
        legend = [
            ("A", "Road Start", "path"),
            ("Z", "Road End", "path"),
            ("P", "Road", "path"),
            ("S", "Slot", "slot"),
            ("B", "Road+Slot", "both"),
            (".", "Empty", "empty"),
        ]
        for offset, (token, label, fill_key) in enumerate(legend, start=3):
            token_cell = ws.cell(row=offset, column=12, value=token)
            token_cell.fill = fills[fill_key]
            token_cell.font = font
            token_cell.alignment = Alignment(horizontal="center", vertical="center")
            token_cell.border = border
            label_cell = ws.cell(row=offset, column=13, value=label)
            label_cell.font = axis_font

    return wb


def main() -> None:
    stage_layouts = parse_stage_layouts()
    workbook = build_workbook(stage_layouts)
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()

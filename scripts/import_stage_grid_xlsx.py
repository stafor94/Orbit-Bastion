from __future__ import annotations

import ast
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "stage_grids_with_start_end.xlsx"
TARGET_PATH = ROOT / "src" / "data" / "slotLayouts.js"
GRID_ROWS = 12
GRID_COLS = 9
ROAD_TOKENS = {"A", "Z", "P", "B"}
SLOT_TOKENS = {"S", "B"}


def neighbors(point: tuple[int, int]) -> list[tuple[int, int]]:
    x, y = point
    return [(x - 1, y), (x + 1, y), (x, y - 1), (x, y + 1)]


def compress_path(path_cells: list[tuple[int, int]]) -> list[list[int]]:
    if len(path_cells) <= 2:
        return [[x, y] for x, y in path_cells]
    result = [path_cells[0]]
    prev_dx = path_cells[1][0] - path_cells[0][0]
    prev_dy = path_cells[1][1] - path_cells[0][1]
    for index in range(1, len(path_cells) - 1):
        current = path_cells[index]
        nxt = path_cells[index + 1]
        dx = nxt[0] - current[0]
        dy = nxt[1] - current[1]
        if (dx, dy) != (prev_dx, prev_dy):
            result.append(current)
        prev_dx, prev_dy = dx, dy
    result.append(path_cells[-1])
    return [[x, y] for x, y in result]


def read_stage_sheet(ws) -> dict[str, list[list[int]]]:
    road_cells: set[tuple[int, int]] = set()
    slot_cells: list[list[int]] = []
    start: tuple[int, int] | None = None
    end: tuple[int, int] | None = None

    for row in range(2, GRID_ROWS + 2):
        for col in range(2, GRID_COLS + 2):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            token = str(value).strip().upper()
            point = (col - 2, row - 2)
            if token in ROAD_TOKENS:
                road_cells.add(point)
            if token in SLOT_TOKENS:
                slot_cells.append([point[0], point[1]])
            if token == "A":
                if start is not None:
                    raise ValueError(f"{ws.title}: multiple A cells")
                start = point
            if token == "Z":
                if end is not None:
                    raise ValueError(f"{ws.title}: multiple Z cells")
                end = point

    if start is None or end is None:
        raise ValueError(f"{ws.title}: missing A or Z")

    adjacency = {point: [n for n in neighbors(point) if n in road_cells] for point in road_cells}
    for point, linked in adjacency.items():
        expected = 1 if point in {start, end} else 2
        if len(linked) != expected:
            raise ValueError(f"{ws.title}: invalid road degree at {point}, expected {expected}, got {len(linked)}")

    ordered = [start]
    previous = None
    current = start
    while current != end:
        options = [point for point in adjacency[current] if point != previous]
        if len(options) != 1:
            raise ValueError(f"{ws.title}: ambiguous path at {current}")
        previous, current = current, options[0]
        ordered.append(current)

    if len(ordered) != len(road_cells):
        raise ValueError(f"{ws.title}: disconnected road cells")

    return {
        "path": compress_path(ordered),
        "slots": slot_cells,
    }


def load_stage_layouts(path: Path) -> list[dict[str, list[list[int]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    layouts: list[dict[str, list[list[int]]]] = []
    for sheet_name in wb.sheetnames:
        layouts.append(read_stage_sheet(wb[sheet_name]))
    return layouts


def update_slot_layouts(path: Path, layouts: list[dict[str, list[list[int]]]]) -> None:
    source = path.read_text(encoding="utf-8")
    stage_lines = ["const STAGE_LAYOUTS = ["]
    for layout in layouts:
        stage_lines.append("    {")
        stage_lines.append(f"      path: {layout['path']},")
        stage_lines.append(f"      slots: {layout['slots']},")
        stage_lines.append("    },")
    stage_lines.append("  ];")
    replacement = "\n".join(stage_lines)
    source = re.sub(r"const STAGE_LAYOUTS = \[(.*?)\];", replacement, source, count=1, flags=re.S)
    path.write_text(source, encoding="utf-8")


def main() -> None:
    workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else WORKBOOK_PATH
    layouts = load_stage_layouts(workbook)
    update_slot_layouts(TARGET_PATH, layouts)
    print(TARGET_PATH)


if __name__ == "__main__":
    main()
